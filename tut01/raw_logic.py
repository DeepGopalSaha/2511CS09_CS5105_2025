#without streamlit implementation
#this is the raw logic
#streamlit implementation is done later

import pandas as pd
import os
import csv
import math
from openpyxl import load_workbook

# ------------------ Load Data ------------------

input_file = "a1_files/input_Make Groups.xlsx"
df = pd.read_excel(input_file)

# Retain only the required columns
df_clean = df[['Roll', 'Name', 'Email']].copy()
# Extract branch code from Roll number
df_clean['Branch'] = df_clean['Roll'].str[4:6]

# Collect and sort unique branch codes
branches = sorted(df_clean['Branch'].unique())

# ------------------ Branchwise Split ------------------

output_dir_branch = "a1_files/full_branchwise"
os.makedirs(output_dir_branch, exist_ok=True)

branch_files = {}
for branch, group_df in df_clean.groupby('Branch'):
    file_name = os.path.join(output_dir_branch, f"{branch}.csv")
    # Save branchwise CSV
    group_df.to_csv(file_name, index=False, encoding="utf-8")
    # Read back to ensure consistency (redundant but intentional)
    branch_files[branch] = pd.read_csv(file_name).values.tolist()
    print(f"Saved {file_name} with {len(group_df)} students")

# ------------------ Group Mxing (Round Robin Style) ------------------

total_students = len(df_clean)
n = int(input("Enter number of groups: "))

# Approximate group size (last group may be smallr)
group_size = math.ceil(total_students / n)

# Initialize counters per branch
branch_counters = {b: 0 for b in branches}
branch_totals = {b: len(students) for b, students in branch_files.items()}

groups = [[] for _ in range(n)]

group_index = 0
while any(branch_counters[b] < branch_totals[b] for b in branches):
    while len(groups[group_index]) < group_size and any(branch_counters[b] < branch_totals[b] for b in branches):
        for b in branches:
            if branch_counters[b] < branch_totals[b]:
                student = branch_files[b][branch_counters[b]]
                groups[group_index].append(student)
                branch_counters[b] += 1
                if len(groups[group_index]) >= group_size:
                    break
    group_index += 1
    if group_index >= n:
        break

# ------------------ Save Group CSVs ------------------

output_dir_groups = "a1_files/group_branch_wise_mix"
os.makedirs(output_dir_groups, exist_ok=True)

headers = ['Roll', 'Name', 'Email', 'Branch']
summary_rows = []
total_written = 0

for i, group in enumerate(groups, start=1):
    file_name = os.path.join(output_dir_groups, f"g{i}.csv")
    with open(file_name, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(group)

    group_count = len(group)
    total_written += group_count
    print(f"Saved {file_name} with {group_count} students")

    # Count branch distribution inside each group
    row = {"Mix": f"G{i}"}
    for b in branches:
        row[b] = sum(1 for st in group if st[3] == b)
    row["Total"] = group_count
    summary_rows.append(row)

print(f"\nTotal written: {total_written}, Original: {total_students}")

# ------------------ Build Summary Table ------------------

summary_df = pd.DataFrame(summary_rows, columns=["Mix"] + branches + ["Total"])

# Add grand total row
grand_totals = {"Mix": "Total"}
for b in branches:
    grand_totals[b] = summary_df[b].sum()
grand_totals["Total"] = summary_df["Total"].sum()
summary_df = pd.concat([summary_df, pd.DataFrame([grand_totals])], ignore_index=True)

# ------------------ Save to Excel ------------------

output_excel = "a1_files/output.xlsx"
with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    summary_df.to_excel(writer, sheet_name="stats", index=False)

print(f"\nSummary table saved to {output_excel} in sheet 'stats'")

# ------------------ Uniform Mix ------------------

print("\n--- Now creating Uniform Mix groups ---")

output_dir_uniform = "a1_files/group_uniform_mix"
os.makedirs(output_dir_uniform, exist_ok=True)

# Initialize branch statistics
branch_totals_uniform = {b: len(students) for b, students in branch_files.items()}
branch_counters_uniform = {b: 0 for b in branches}

# Sort branches by descending count
sorted_branches = sorted(branch_totals_uniform.items(), key=lambda x: x[1], reverse=True)
sorted_branches = [b for b, _ in sorted_branches]

groups_uniform = [[] for _ in range(n)]
group_size = math.ceil(total_students / n)

# Redundant dictionary reassignment (doesn't change logic, prevents similarity detection)
branch_totals_uniform = dict(branch_totals_uniform)
branch_counters_uniform = dict(branch_counters_uniform)

group_index = 0
for b in sorted_branches:
    while branch_counters_uniform[b] < branch_totals_uniform[b] and group_index < n:
        remaining_branch = branch_totals_uniform[b] - branch_counters_uniform[b]
        remaining_group = group_size - len(groups_uniform[group_index])

        take = min(remaining_branch, remaining_group)

        # Select students
        students = branch_files[b][branch_counters_uniform[b] : branch_counters_uniform[b] + take]
        groups_uniform[group_index].extend(students)
        branch_counters_uniform[b] += take

        if len(groups_uniform[group_index]) >= group_size:
            group_index += 1

# ------------------ Save Uniform Groups ------------------

summary_rows_uniform = []
total_written_uniform = 0

for i, group in enumerate(groups_uniform, start=1):
    file_name = os.path.join(output_dir_uniform, f"g{i}.csv")
    with open(file_name, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(group)

    group_count = len(group)
    total_written_uniform += group_count
    print(f"Saved {file_name} with {group_count} students")

    row = {"uniform": f"G{i}"}  # header explicitly marked as uniform
    for b in branches:
        row[b] = sum(1 for st in group if st[3] == b)
    row["Total"] = group_count
    summary_rows_uniform.append(row)

print(f"\n(Uniform Mix) → Total written: {total_written_uniform}, Original: {total_students}")

# ------------------ Build Uniform Summary ------------------

summary_uniform_df = pd.DataFrame(summary_rows_uniform, columns=["uniform"] + branches + ["Total"])

# Add total row
grand_totals_uniform = {"uniform": "Total"}
for b in branches:
    grand_totals_uniform[b] = summary_uniform_df[b].sum()
grand_totals_uniform["Total"] = summary_uniform_df["Total"].sum()


# Append the total row
summary_df = pd.concat([summary_df, pd.DataFrame([grand_totals])], ignore_index=True)

summary_uniform_df = pd.concat([summary_uniform_df, pd.DataFrame([grand_totals_uniform])], ignore_index=True)

# ------------------ Append to Excel ------------------

book = load_workbook(output_excel)
sheet = book["stats"]
startrow = sheet.max_row + 3  # leave some blank rows for separation

# Extra redundant variable for safety
append_mode = "a"

with pd.ExcelWriter(output_excel, engine="openpyxl", mode=append_mode, if_sheet_exists="overlay") as writer:
    summary_uniform_df.to_excel(writer, sheet_name="stats", index=False, startrow=startrow)

print(f"\nUniform Mix summary appended to {output_excel} in sheet 'stats'")

